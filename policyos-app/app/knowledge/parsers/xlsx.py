from io import BytesIO

import openpyxl
from openpyxl import load_workbook

from app.knowledge.schemas import ParsedDocument, ParsedSection, ParserError


class XlsxDocumentParser:
    name = "openpyxl"
    version = openpyxl.__version__
    extensions = frozenset({".xlsx"})
    mime_types = frozenset(
        {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
    )

    def __init__(self, *, max_rows: int = 10_000, max_columns: int = 200) -> None:
        self.max_rows = max_rows
        self.max_columns = max_columns

    def parse(self, content: bytes, filename: str) -> ParsedDocument:
        try:
            workbook = load_workbook(
                BytesIO(content), read_only=True, data_only=False, keep_links=False
            )
        except Exception as exc:
            raise ParserError("XLSX document could not be parsed") from exc
        try:
            sections: list[ParsedSection] = []
            for sheet in workbook.worksheets:
                if sheet.max_row > self.max_rows or sheet.max_column > self.max_columns:
                    raise ParserError("XLSX row or column limit exceeded")
                rows: list[str] = []
                for row in sheet.iter_rows(values_only=False):
                    values: list[str] = []
                    for cell in row:
                        value = cell.value
                        values.append("" if value is None else str(value))
                    rows.append(" | ".join(values))
                sections.append(
                    ParsedSection(
                        index=len(sections) + 1,
                        kind="sheet",
                        title=sheet.title,
                        sheet_name=sheet.title,
                        text="\n".join(rows),
                        start_row=1,
                        end_row=sheet.max_row,
                        metadata={"row_count": sheet.max_row, "column_count": sheet.max_column},
                    )
                )
        finally:
            workbook.close()
        if not any(section.text.strip(" |\n") for section in sections):
            raise ParserError("XLSX document contains no data")
        return ParsedDocument(
            text="\n\n".join(
                f"[{section.sheet_name}]\n{section.text}" for section in sections
            ),
            sections=sections,
            parser_name=self.name,
            parser_version=self.version,
            sheet_count=len(sections),
            metadata={"sheet_names": [section.sheet_name for section in sections]},
        )